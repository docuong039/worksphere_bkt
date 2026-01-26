import openpyxl

excel_path = r'c:\Users\Admin\Documents\worksphere_bkt\docs\UI_STORY_MATRIX_UPDATED.xlsx'
wb = openpyxl.load_workbook(excel_path)

mapping = {
    '/admin/custom-fields': [
        ('4. SYS_ADMIN', 'US-SYS-01-05'), 
        ('5. ORG_ADMIN', 'US-ORG-03-03')
    ],
    '/projects/[id]/activity': [
        ('1. EMP', 'US-EMP-04-01'), 
        ('1. EMP', 'US-EMP-04-02'),
        ('2. PM (MNG)', 'US-MNG-06-01'), 
        ('2. PM (MNG)', 'US-MNG-06-02'), 
        ('3. CEO', 'US-CEO-04-01'),
        ('3. CEO', 'US-CEO-04-02')
    ],
    '/projects/[id]/settings': [
        ('2. PM (MNG)', 'US-MNG-01-01'),
        ('2. PM (MNG)', 'US-MNG-01-13') # Phân quyền trường (often in settings)
    ],
    '/projects/[id]/settings/lockdown': [
        ('2. PM (MNG)', 'US-MNG-04-01'), 
        ('2. PM (MNG)', 'US-MNG-04-02')
    ],
    '/projects/[id]/settings/recycle-bin': [
        ('2. PM (MNG)', 'US-MNG-08-01'), 
        ('2. PM (MNG)', 'US-MNG-08-02'),
        ('2. PM (MNG)', 'US-MNG-08-03')
    ],
    '/projects/[id]/time-logs': [
        ('1. EMP', 'US-EMP-02-03'), 
        ('2. PM (MNG)', 'US-MNG-04-01'),
        ('2. PM (MNG)', 'US-MNG-04-02'),
        ('2. PM (MNG)', 'US-MNG-07-01')
    ],
    '/register-org': [
        ('4. SYS_ADMIN', 'US-SYS-01-01')
    ]
}

def clean(s):
    if not s: return ""
    return str(s).strip()

for route, assignments in mapping.items():
    for sheet_name, usid in assignments:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        route_col = 2
        
        # Find USID Row
        usid_row = 3
        for r in range(1, 10):
            val = clean(ws.cell(row=r, column=3).value)
            if val.startswith('US-'):
                usid_row = r
                break
        
        # Find USID Column
        target_col = None
        for c in range(3, ws.max_column + 1):
            if clean(ws.cell(row=usid_row, column=c).value) == usid:
                target_col = c
                break
        
        # Find Route Row
        target_row = None
        for r in range(usid_row + 1, ws.max_row + 1):
            if clean(ws.cell(row=r, column=route_col).value) == route:
                target_row = r
                break
        
        if target_row and target_col:
            ws.cell(row=target_row, column=target_col).value = 'X'
            print(f"Mapped {route} to {usid} in {sheet_name}")

wb.save(excel_path)
print("Excel mapping updated.")
