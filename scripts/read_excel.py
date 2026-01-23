
import os
import sys
from openpyxl import load_workbook

# Set encoding to utf-8 for console output
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILE_PATH = os.path.join(BASE_DIR, "docs", "UI_STORY_MATRIX.xlsx")

def read_full_excel_to_md():
    if not os.path.exists(FILE_PATH):
        print(f"File not found: {FILE_PATH}")
        return

    try:
        wb = load_workbook(FILE_PATH, data_only=True)
        output_file = "UI_STORY_MATRIX.md"
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(f"# UI-STORY MATRIX CHECKLIST\n\n")
            f.write(f"**Source:** `{FILE_PATH}`\n\n")
            
            for sheet_name in wb.sheetnames:
                f.write(f"## SHEET: {sheet_name}\n\n")
                ws = wb[sheet_name]
                
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                
                # Filter out completely empty rows
                filtered_rows = []
                for row in rows:
                    cleaned_row = [str(cell).replace('\n', ' ').strip() if cell is not None else "" for cell in row]
                    if any(cleaned_row):
                        filtered_rows.append(cleaned_row)
                
                if not filtered_rows:
                    continue

                # Process as a Markdown table
                # We'll assume the first row is the header
                header = filtered_rows[0]
                f.write("| " + " | ".join(header) + " |\n")
                f.write("| " + " | ".join(["---"] * len(header)) + " |\n")
                
                for row in filtered_rows[1:]:
                    f.write("| " + " | ".join(row) + " |\n")
                
                f.write("\n---\n\n")
                        
        print(f"Successfully converted Excel to Markdown: {output_file}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    read_full_excel_to_md()
