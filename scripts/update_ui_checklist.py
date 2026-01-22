#!/usr/bin/env python3
"""
Script to update UI_CHECKLIST_FULL.xlsx with current source code state.
Aligns with 5 BA documentation files and current page.tsx files.
"""

import os
import re
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(BASE_DIR, "src", "app", "(frontend)")
OUTPUT_FILE = os.path.join(BASE_DIR, "docs", "UI_CHECKLIST_UPDATED.xlsx")

# User Story mappings from BA documentation
USER_STORY_MAP = {
    # Authentication
    "/login": "US-EMP-00-01, US-MNG-00-01, US-CEO-00-01, US-SYS-00-01, US-ORG-00-01",
    "/join": "US-ORG-01-02",
    "/forgot-password": "US-EMP-00-03, US-MNG-00-03, US-CEO-00-03",
    "/reset-password": "US-EMP-00-03",
    
    # Dashboard
    "/dashboard": "US-EMP-01-01, US-MNG-02-01, US-CEO-01-01",
    "/executive/dashboard": "US-CEO-01-01, US-CEO-01-02, US-CEO-01-03",
    
    # Tasks
    "/tasks": "US-EMP-01-01, US-EMP-01-02",
    "/tasks/[id]": "US-EMP-01-03, US-EMP-01-04, US-EMP-01-05, US-EMP-01-06, US-EMP-01-07, US-EMP-01-08, US-EMP-01-09, US-EMP-01-10, US-EMP-01-11",
    "/tasks/new": "US-PM-01-02",
    "/tasks/kanban": "US-EMP-01-01, US-CEO-07-01",
    
    # Projects
    "/projects": "US-PM-01-01",
    "/projects/new": "US-PM-01-01",
    "/projects/[id]": "US-PM-01-02",
    "/projects/[id]/overview": "US-PM-01-03, US-PM-02-01, US-PM-02-02",
    "/projects/[id]/gantt": "US-PM-02-03",
    "/projects/[id]/resources": "US-PM-03-01, US-PM-03-02",
    "/projects/[id]/documents": "US-PM-03-03",
    "/projects/[id]/import-export": "US-PM-04-01, US-PM-04-02",
    "/projects/[id]/time-locks": "US-PM-05-01, US-PM-05-02",
    "/projects/[id]/cost": "US-MNG-03-01, US-CEO-02-02",
    "/projects/[id]/performance": "US-CEO-01-03",
    "/projects/[id]/settings/custom-fields": "US-PM-06-01",
    "/projects/[id]/settings/field-permissions": "US-PM-06-02",
    "/projects/[id]/settings/tags": "US-PM-06-03",
    "/projects/[id]/settings/notifications": "US-PM-06-04",
    
    # Reports
    "/reports": "US-MNG-02-02, US-CEO-01-04",
    "/reports/new": "US-MNG-02-02",
    "/reports/[id]": "US-MNG-02-02",
    "/reports/cost-analysis": "US-MNG-03-01, US-CEO-02-02",
    "/reports/performance": "US-CEO-01-03",
    
    # HR
    "/hr/employees": "US-ORG-01-01, US-ORG-01-02, US-ORG-01-03, US-ORG-01-04, US-ORG-01-05",
    "/hr/employees/[id]": "US-CEO-01-02, US-ORG-01-03",
    "/hr/employees/[id]/timeline": "US-CEO-01-02",
    "/hr/salary": "US-MNG-03-02, US-CEO-02-01",
    "/hr/contracts": "US-ORG-02-01, US-ORG-02-02",
    
    # Time Logs
    "/time-logs": "US-EMP-02-01, US-EMP-02-02, US-EMP-02-03",
    
    # Activity & Notifications
    "/activity": "US-EMP-03-01",
    "/notifications": "US-EMP-03-02",
    "/personal-board": "US-EMP-01-01, US-CEO-07-01",
    "/recycle-bin": "US-PM-07-01, US-ORG-03-01",
    
    # Settings
    "/settings/profile": "US-EMP-00-02",
    "/settings/workspace": "US-ORG-04-01",
    "/settings/lookups": "US-ORG-04-02",
    "/settings/auto-lock": "US-ORG-04-03",
    "/settings/recycle-bin": "US-ORG-03-01",
    
    # Admin
    "/admin/users": "US-SYS-01-01, US-SYS-01-02, US-SYS-01-03",
    "/admin/users/[id]": "US-SYS-01-02, US-SYS-01-03",
    "/admin/organizations": "US-SYS-02-01, US-SYS-02-02",
    "/admin/org-approvals": "US-SYS-02-03",
    "/admin/deleted-orgs": "US-SYS-02-04",
    "/admin/org-recycle-bin": "US-SYS-02-04",
    "/admin/members": "US-ORG-01-01, US-ORG-01-02",
    "/admin/invites": "US-ORG-01-02",
    "/admin/roles": "US-SYS-03-01, US-ORG-05-01",
    "/admin/roles/customize": "US-SYS-03-02, US-ORG-05-02",
    "/admin/quotas": "US-SYS-04-01",
    "/admin/audit-logs": "US-SYS-05-01",
    "/admin/platform-dashboard": "US-SYS-02-01",
    "/admin/impersonation": "US-SYS-06-01",
}

# Database tables mapping
DATABASE_MAP = {
    "/login": "users, organizations",
    "/join": "users, user_invitations, organizations",
    "/forgot-password": "users, password_reset_tokens",
    "/reset-password": "users, password_reset_tokens",
    "/dashboard": "tasks, projects, time_logs, activity_events",
    "/executive/dashboard": "projects, tasks, users, organizations",
    "/tasks": "tasks, subtasks, task_comments, task_attachments",
    "/tasks/[id]": "tasks, subtasks, task_comments, task_attachments, time_logs",
    "/tasks/new": "tasks, projects, users",
    "/tasks/kanban": "tasks",
    "/projects": "projects, project_members",
    "/projects/new": "projects",
    "/projects/[id]": "projects, tasks, project_members",
    "/projects/[id]/overview": "projects, tasks, project_members, time_logs",
    "/projects/[id]/gantt": "tasks, subtasks",
    "/projects/[id]/resources": "project_resources",
    "/projects/[id]/documents": "project_documents, task_attachments",
    "/projects/[id]/import-export": "tasks, subtasks",
    "/projects/[id]/time-locks": "time_log_locks",
    "/projects/[id]/cost": "time_logs, user_compensations",
    "/projects/[id]/performance": "tasks, time_logs",
    "/projects/[id]/settings/custom-fields": "custom_field_definitions",
    "/projects/[id]/settings/field-permissions": "custom_field_definitions, roles",
    "/projects/[id]/settings/tags": "tags",
    "/projects/[id]/settings/notifications": "notification_settings",
    "/reports": "reports",
    "/reports/new": "reports",
    "/reports/[id]": "reports",
    "/reports/cost-analysis": "time_logs, user_compensations, projects",
    "/reports/performance": "tasks, time_logs, users",
    "/hr/employees": "users, organizations",
    "/hr/employees/[id]": "users, user_compensations, projects",
    "/hr/employees/[id]/timeline": "users, user_compensations, project_members",
    "/hr/salary": "job_levels, user_compensations",
    "/hr/contracts": "contracts, users",
    "/time-logs": "time_logs, tasks, projects",
    "/activity": "activity_events",
    "/notifications": "notifications",
    "/personal-board": "tasks",
    "/recycle-bin": "soft_deleted_items",
    "/settings/profile": "users",
    "/settings/workspace": "organizations",
    "/settings/lookups": "lookup_values",
    "/settings/auto-lock": "auto_lock_settings",
    "/settings/recycle-bin": "soft_deleted_items",
    "/admin/users": "users, roles",
    "/admin/users/[id]": "users, roles",
    "/admin/organizations": "organizations",
    "/admin/org-approvals": "organizations",
    "/admin/deleted-orgs": "organizations",
    "/admin/org-recycle-bin": "organizations",
    "/admin/members": "users, organizations",
    "/admin/invites": "user_invitations",
    "/admin/roles": "roles, permissions",
    "/admin/roles/customize": "roles, permissions, role_permissions",
    "/admin/quotas": "quotas, organizations",
    "/admin/audit-logs": "audit_logs",
    "/admin/platform-dashboard": "organizations, users, projects",
    "/admin/impersonation": "users, audit_logs",
}

# Role access mapping
ROLE_MAP = {
    "/login": "ALL",
    "/join": "PUBLIC",
    "/forgot-password": "PUBLIC",
    "/reset-password": "PUBLIC",
    "/dashboard": "EMP, PM, CEO, ORG_ADMIN",
    "/executive/dashboard": "CEO",
    "/tasks": "EMP, PM, CEO",
    "/tasks/[id]": "EMP, PM, CEO",
    "/tasks/new": "PM",
    "/tasks/kanban": "EMP, PM, CEO",
    "/projects": "PM, CEO, ORG_ADMIN",
    "/projects/new": "PM, ORG_ADMIN",
    "/projects/[id]": "PM, EMP",
    "/projects/[id]/overview": "PM, EMP, CEO",
    "/projects/[id]/gantt": "PM, EMP",
    "/projects/[id]/resources": "PM",
    "/projects/[id]/documents": "PM, EMP",
    "/projects/[id]/import-export": "PM",
    "/projects/[id]/time-locks": "PM",
    "/projects/[id]/cost": "PM, CEO",
    "/projects/[id]/performance": "PM, CEO",
    "/projects/[id]/settings/custom-fields": "PM",
    "/projects/[id]/settings/field-permissions": "PM",
    "/projects/[id]/settings/tags": "PM",
    "/projects/[id]/settings/notifications": "PM",
    "/reports": "PM, CEO, ORG_ADMIN",
    "/reports/new": "PM, CEO",
    "/reports/[id]": "PM, CEO",
    "/reports/cost-analysis": "PM, CEO",
    "/reports/performance": "CEO",
    "/hr/employees": "ORG_ADMIN, CEO",
    "/hr/employees/[id]": "ORG_ADMIN, CEO",
    "/hr/employees/[id]/timeline": "CEO",
    "/hr/salary": "ORG_ADMIN, CEO",
    "/hr/contracts": "ORG_ADMIN, CEO",
    "/time-logs": "EMP, PM, CEO",
    "/activity": "EMP, PM, CEO",
    "/notifications": "EMP, PM, CEO, ORG_ADMIN",
    "/personal-board": "EMP, PM, CEO",
    "/recycle-bin": "PM, ORG_ADMIN",
    "/settings/profile": "EMP, PM, CEO, ORG_ADMIN",
    "/settings/workspace": "ORG_ADMIN",
    "/settings/lookups": "ORG_ADMIN",
    "/settings/auto-lock": "ORG_ADMIN",
    "/settings/recycle-bin": "ORG_ADMIN",
    "/admin/users": "SYS_ADMIN",
    "/admin/users/[id]": "SYS_ADMIN",
    "/admin/organizations": "SYS_ADMIN",
    "/admin/org-approvals": "SYS_ADMIN",
    "/admin/deleted-orgs": "SYS_ADMIN",
    "/admin/org-recycle-bin": "SYS_ADMIN",
    "/admin/members": "ORG_ADMIN",
    "/admin/invites": "ORG_ADMIN",
    "/admin/roles": "SYS_ADMIN, ORG_ADMIN",
    "/admin/roles/customize": "SYS_ADMIN, ORG_ADMIN",
    "/admin/quotas": "SYS_ADMIN",
    "/admin/audit-logs": "SYS_ADMIN",
    "/admin/platform-dashboard": "SYS_ADMIN",
    "/admin/impersonation": "SYS_ADMIN",
}

# UI Name mapping
UI_NAME_MAP = {
    "/login": "Đăng nhập",
    "/join": "Tham gia tổ chức",
    "/forgot-password": "Quên mật khẩu",
    "/reset-password": "Đặt lại mật khẩu",
    "/dashboard": "Bảng điều khiển",
    "/executive/dashboard": "Dashboard Giám đốc",
    "/tasks": "Danh sách công việc",
    "/tasks/[id]": "Chi tiết công việc",
    "/tasks/new": "Tạo công việc mới",
    "/tasks/kanban": "Kanban công việc",
    "/projects": "Danh sách dự án",
    "/projects/new": "Tạo dự án mới",
    "/projects/[id]": "Chi tiết dự án - Tasks",
    "/projects/[id]/overview": "Tổng quan dự án",
    "/projects/[id]/gantt": "Biểu đồ Gantt",
    "/projects/[id]/resources": "Tài nguyên dự án",
    "/projects/[id]/documents": "Tài liệu dự án",
    "/projects/[id]/import-export": "Import/Export dự án",
    "/projects/[id]/time-locks": "Khóa kỳ công",
    "/projects/[id]/cost": "Chi phí dự án",
    "/projects/[id]/performance": "Hiệu suất dự án",
    "/projects/[id]/settings/custom-fields": "Trường tùy chỉnh",
    "/projects/[id]/settings/field-permissions": "Phân quyền trường",
    "/projects/[id]/settings/tags": "Quản lý nhãn",
    "/projects/[id]/settings/notifications": "Cài đặt thông báo",
    "/reports": "Danh sách báo cáo",
    "/reports/new": "Tạo báo cáo mới",
    "/reports/[id]": "Chi tiết báo cáo",
    "/reports/cost-analysis": "Phân tích chi phí",
    "/reports/performance": "Báo cáo hiệu suất",
    "/hr/employees": "Danh sách nhân sự",
    "/hr/employees/[id]": "Hồ sơ nhân viên",
    "/hr/employees/[id]/timeline": "Lịch sử làm việc",
    "/hr/salary": "Quản lý lương & cấp bậc",
    "/hr/contracts": "Quản lý hợp đồng",
    "/time-logs": "Nhật ký thời gian",
    "/activity": "Hoạt động gần đây",
    "/notifications": "Thông báo",
    "/personal-board": "Bảng cá nhân",
    "/recycle-bin": "Thùng rác",
    "/settings/profile": "Hồ sơ cá nhân",
    "/settings/workspace": "Cài đặt Workspace",
    "/settings/lookups": "Danh mục hệ thống",
    "/settings/auto-lock": "Khóa tự động",
    "/settings/recycle-bin": "Thùng rác Settings",
    "/admin/users": "Quản lý người dùng",
    "/admin/users/[id]": "Chi tiết người dùng",
    "/admin/organizations": "Quản lý tổ chức",
    "/admin/org-approvals": "Duyệt tổ chức",
    "/admin/deleted-orgs": "Tổ chức đã xóa",
    "/admin/org-recycle-bin": "Thùng rác tổ chức",
    "/admin/members": "Thành viên tổ chức",
    "/admin/invites": "Lời mời đang chờ",
    "/admin/roles": "Quản lý vai trò",
    "/admin/roles/customize": "Tùy chỉnh quyền",
    "/admin/quotas": "Hạn mức hệ thống",
    "/admin/audit-logs": "Nhật ký kiểm toán",
    "/admin/platform-dashboard": "Dashboard Platform",
    "/admin/impersonation": "Impersonation",
}

# Group mapping
GROUP_MAP = {
    "/login": "Authentication",
    "/join": "Authentication",
    "/forgot-password": "Authentication",
    "/reset-password": "Authentication",
    "/dashboard": "Dashboard",
    "/executive/dashboard": "Dashboard",
    "/tasks": "Tasks",
    "/tasks/[id]": "Tasks",
    "/tasks/new": "Tasks",
    "/tasks/kanban": "Tasks",
    "/projects": "Projects",
    "/projects/new": "Projects",
    "/projects/[id]": "Projects",
    "/projects/[id]/overview": "Project Detail",
    "/projects/[id]/gantt": "Project Detail",
    "/projects/[id]/resources": "Project Detail",
    "/projects/[id]/documents": "Project Detail",
    "/projects/[id]/import-export": "Project Detail",
    "/projects/[id]/time-locks": "Project Detail",
    "/projects/[id]/cost": "Project Detail",
    "/projects/[id]/performance": "Project Detail",
    "/projects/[id]/settings/custom-fields": "Project Settings",
    "/projects/[id]/settings/field-permissions": "Project Settings",
    "/projects/[id]/settings/tags": "Project Settings",
    "/projects/[id]/settings/notifications": "Project Settings",
    "/reports": "Reports",
    "/reports/new": "Reports",
    "/reports/[id]": "Reports",
    "/reports/cost-analysis": "Reports",
    "/reports/performance": "Reports",
    "/hr/employees": "HR Management",
    "/hr/employees/[id]": "HR Management",
    "/hr/employees/[id]/timeline": "HR Management",
    "/hr/salary": "HR Management",
    "/hr/contracts": "HR Management",
    "/time-logs": "Time Management",
    "/activity": "Activity",
    "/notifications": "Notifications",
    "/personal-board": "Personal",
    "/recycle-bin": "System",
    "/settings/profile": "Settings",
    "/settings/workspace": "Settings",
    "/settings/lookups": "Settings",
    "/settings/auto-lock": "Settings",
    "/settings/recycle-bin": "Settings",
    "/admin/users": "Admin - System",
    "/admin/users/[id]": "Admin - System",
    "/admin/organizations": "Admin - System",
    "/admin/org-approvals": "Admin - System",
    "/admin/deleted-orgs": "Admin - System",
    "/admin/org-recycle-bin": "Admin - System",
    "/admin/members": "Admin - Organization",
    "/admin/invites": "Admin - Organization",
    "/admin/roles": "Admin - RBAC",
    "/admin/roles/customize": "Admin - RBAC",
    "/admin/quotas": "Admin - System",
    "/admin/audit-logs": "Admin - System",
    "/admin/platform-dashboard": "Admin - System",
    "/admin/impersonation": "Admin - System",
}

# Function descriptions
FUNCTION_MAP = {
    "/login": "Đăng nhập vào hệ thống với email và mật khẩu",
    "/join": "Tham gia tổ chức qua link mời",
    "/forgot-password": "Yêu cầu link đặt lại mật khẩu qua email",
    "/reset-password": "Đặt lại mật khẩu mới từ link email",
    "/dashboard": "Xem tổng quan công việc, thống kê, hoạt động gần đây",
    "/executive/dashboard": "Xem tổng quan cấp cao: dự án, chi phí, nhân sự",
    "/tasks": "Xem danh sách công việc được giao, lọc, tìm kiếm",
    "/tasks/[id]": "Xem chi tiết, cập nhật trạng thái, log time, comment",
    "/tasks/new": "Tạo công việc mới với đầy đủ thông tin",
    "/tasks/kanban": "Quản lý công việc dạng Kanban, kéo thả trạng thái",
    "/projects": "Xem danh sách dự án, lọc theo trạng thái",
    "/projects/new": "Tạo dự án mới với thông tin cơ bản",
    "/projects/[id]": "Xem danh sách tasks trong dự án",
    "/projects/[id]/overview": "Tổng quan dự án: tiến độ, thành viên, thống kê",
    "/projects/[id]/gantt": "Xem và chỉnh sửa timeline dạng Gantt chart",
    "/projects/[id]/resources": "Quản lý tài nguyên: links, documents, Git repos",
    "/projects/[id]/documents": "Xem và quản lý tài liệu dự án",
    "/projects/[id]/import-export": "Import/Export tasks từ Excel, CSV",
    "/projects/[id]/time-locks": "Khóa/mở khóa kỳ chấm công",
    "/projects/[id]/cost": "Xem báo cáo chi phí dự án",
    "/projects/[id]/performance": "Xem hiệu suất dự án",
    "/projects/[id]/settings/custom-fields": "Tạo và quản lý trường tùy chỉnh",
    "/projects/[id]/settings/field-permissions": "Phân quyền truy cập trường dữ liệu",
    "/projects/[id]/settings/tags": "Quản lý nhãn/tags cho tasks",
    "/projects/[id]/settings/notifications": "Cài đặt thông báo dự án",
    "/reports": "Xem danh sách báo cáo đã tạo",
    "/reports/new": "Tạo báo cáo mới",
    "/reports/[id]": "Xem chi tiết báo cáo",
    "/reports/cost-analysis": "Phân tích chi phí theo dự án, nhân sự",
    "/reports/performance": "Báo cáo hiệu suất đội ngũ",
    "/hr/employees": "Quản lý danh sách nhân sự, mời thành viên mới",
    "/hr/employees/[id]": "Xem hồ sơ chi tiết nhân viên",
    "/hr/employees/[id]/timeline": "Xem lịch sử thăng tiến, thay đổi lương",
    "/hr/salary": "Quản lý cấp bậc và mức lương nhân sự",
    "/hr/contracts": "Quản lý hợp đồng lao động",
    "/time-logs": "Xem và quản lý nhật ký thời gian làm việc",
    "/activity": "Xem lịch sử hoạt động trong hệ thống",
    "/notifications": "Xem và quản lý thông báo",
    "/personal-board": "Bảng công việc cá nhân dạng Kanban",
    "/recycle-bin": "Xem và khôi phục items đã xóa",
    "/settings/profile": "Cập nhật thông tin cá nhân, đổi mật khẩu",
    "/settings/workspace": "Cài đặt tổ chức: tên, logo, timezone",
    "/settings/lookups": "Quản lý danh mục hệ thống",
    "/settings/auto-lock": "Cài đặt tự động khóa kỳ công",
    "/settings/recycle-bin": "Quản lý thùng rác cấp settings",
    "/admin/users": "Quản lý tất cả users trên platform",
    "/admin/users/[id]": "Xem/sửa chi tiết user",
    "/admin/organizations": "Quản lý tất cả organizations",
    "/admin/org-approvals": "Duyệt đơn đăng ký tổ chức mới",
    "/admin/deleted-orgs": "Xem tổ chức đã xóa",
    "/admin/org-recycle-bin": "Khôi phục hoặc xóa vĩnh viễn tổ chức",
    "/admin/members": "Quản lý thành viên trong tổ chức",
    "/admin/invites": "Quản lý lời mời đang chờ",
    "/admin/roles": "Quản lý vai trò và quyền hạn",
    "/admin/roles/customize": "Tùy chỉnh quyền chi tiết cho vai trò",
    "/admin/quotas": "Cài đặt hạn mức sử dụng của tổ chức",
    "/admin/audit-logs": "Xem lịch sử thao tác hệ thống",
    "/admin/platform-dashboard": "Dashboard tổng quan platform",
    "/admin/impersonation": "Đăng nhập thay user để hỗ trợ",
}


def path_to_route(file_path):
    """Convert file path to route"""
    route = file_path.replace("\\", "/")
    route = route.replace("src/app/(frontend)/", "")
    route = route.replace("/page.tsx", "")
    route = route.replace("(auth)/", "")
    if not route.startswith("/"):
        route = "/" + route
    return route


def extract_testids(file_path):
    """Extract all data-testid values from a file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        # Find all data-testid patterns
        pattern = r'data-testid[=]"([^"]+)"'
        matches = re.findall(pattern, content)
        # Also find dynamic testids like data-testid={`xxx-${id}`}
        pattern2 = r'data-testid[=]\{[`\']([^`\']+)[`\']\}'
        matches2 = re.findall(pattern2, content)
        all_matches = list(set(matches + matches2))
        return sorted(all_matches)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []


def get_all_pages():
    """Get all page.tsx files"""
    pages = []
    for root, dirs, files in os.walk(FRONTEND_DIR):
        for file in files:
            if file == "page.tsx":
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, BASE_DIR)
                route = path_to_route(rel_path)
                testids = extract_testids(full_path)
                pages.append({
                    "route": route,
                    "source_file": rel_path.replace("\\", "/"),
                    "testids": testids,
                    "testid_count": len(testids)
                })
    return sorted(pages, key=lambda x: x["route"])


def create_checklist():
    """Create the updated Excel checklist"""
    wb = Workbook()
    
    # Sheet 1: UI Checklist
    ws1 = wb.active
    ws1.title = "1. UI Checklist"
    
    # Headers
    headers = [
        "STT", "Nhóm", "Loại", "Tên UI", "Route/URL", "Chức năng",
        "User Story ID", "Database Tables", "data-testid (Từ Source)",
        "Role", "Source File", "Trạng thái"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Get all pages
    pages = get_all_pages()
    
    # Write data
    for idx, page in enumerate(pages, 1):
        route = page["route"]
        row = idx + 1
        
        ws1.cell(row=row, column=1, value=idx)  # STT
        ws1.cell(row=row, column=2, value=GROUP_MAP.get(route, "Other"))  # Nhóm
        ws1.cell(row=row, column=3, value="Page")  # Loại
        ws1.cell(row=row, column=4, value=UI_NAME_MAP.get(route, route))  # Tên UI
        ws1.cell(row=row, column=5, value=route)  # Route/URL
        ws1.cell(row=row, column=6, value=FUNCTION_MAP.get(route, ""))  # Chức năng
        ws1.cell(row=row, column=7, value=USER_STORY_MAP.get(route, ""))  # User Story ID
        ws1.cell(row=row, column=8, value=DATABASE_MAP.get(route, ""))  # Database Tables
        ws1.cell(row=row, column=9, value=", ".join(page["testids"]) if page["testids"] else "")  # data-testid - ALL
        ws1.cell(row=row, column=10, value=ROLE_MAP.get(route, ""))  # Role
        ws1.cell(row=row, column=11, value=page["source_file"])  # Source File
        ws1.cell(row=row, column=12, value="Done" if page["testid_count"] >= 5 else "Need Review")  # Trạng thái
        
        # Apply borders
        for col in range(1, 13):
            ws1.cell(row=row, column=col).border = thin_border
    
    # Adjust column widths
    column_widths = [5, 18, 8, 25, 35, 50, 40, 40, 80, 25, 50, 15]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width
    
    # Sheet 2: TestId by File
    ws2 = wb.create_sheet("2. TestId by File")
    
    headers2 = ["Source File", "Số lượng", "data-testid List"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for idx, page in enumerate(pages, 1):
        row = idx + 1
        ws2.cell(row=row, column=1, value=page["source_file"])
        ws2.cell(row=row, column=2, value=page["testid_count"])
        ws2.cell(row=row, column=3, value=", ".join(page["testids"]))
        
        for col in range(1, 4):
            ws2.cell(row=row, column=col).border = thin_border
    
    ws2.column_dimensions['A'].width = 60
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 150
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"✅ Updated {OUTPUT_FILE}")
    print(f"   Total pages: {len(pages)}")
    print(f"   Total testids: {sum(p['testid_count'] for p in pages)}")


if __name__ == "__main__":
    create_checklist()
