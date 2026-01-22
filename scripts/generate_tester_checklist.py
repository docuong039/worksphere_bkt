#!/usr/bin/env python3
"""
Script to generate comprehensive UI Checklist Excel for Playwright testing.
- Sheet 1: UI Checklist (overview)
- Sheet 2: TestId by File
- Sheet 3+: Detailed sheets by module with each data-testid on separate row
"""

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(BASE_DIR, "src", "app", "(frontend)")
OUTPUT_FILE = os.path.join(BASE_DIR, "docs", "UI_CHECKLIST_TESTER.xlsx")

# Element type detection patterns
ELEMENT_TYPES = {
    'btn-': 'Button',
    'button': 'Button',
    'submit': 'Button',
    'input-': 'Input',
    'select-': 'Select',
    'filter-': 'Filter',
    'search': 'Input',
    'checkbox': 'Checkbox',
    'toggle': 'Toggle',
    'switch': 'Toggle',
    'table': 'Table',
    'row-': 'TableRow',
    'dialog': 'Dialog',
    'modal': 'Dialog',
    'container': 'Container',
    'page': 'Container',
    'title': 'Heading',
    'header': 'Header',
    'card': 'Card',
    'stat-': 'StatCard',
    'tab': 'Tab',
    'link': 'Link',
    'loading': 'Loading',
    'skeleton': 'Loading',
    'empty': 'EmptyState',
    'error': 'ErrorMessage',
    'alert': 'Alert',
    'form': 'Form',
    'list': 'List',
    'menu': 'Menu',
    'dropdown': 'Dropdown',
    'avatar': 'Avatar',
    'badge': 'Badge',
    'icon': 'Icon',
    'image': 'Image',
    'label': 'Label',
    'text': 'Text',
    'tooltip': 'Tooltip',
    'pagination': 'Pagination',
    'sidebar': 'Sidebar',
    'nav': 'Navigation',
}

# Action suggestions based on element type
ACTION_MAP = {
    'Button': 'Click',
    'Input': 'Fill text',
    'Select': 'Select option',
    'Filter': 'Select filter',
    'Checkbox': 'Check/Uncheck',
    'Toggle': 'Toggle on/off',
    'Table': 'Check visible, count rows',
    'TableRow': 'Click to select',
    'Dialog': 'Check visible, close',
    'Container': 'Check visible',
    'Heading': 'Get text, check visible',
    'Header': 'Check visible',
    'Card': 'Check visible, click',
    'StatCard': 'Get value',
    'Tab': 'Click to switch',
    'Link': 'Click to navigate',
    'Loading': 'Wait for disappear',
    'EmptyState': 'Check visible when no data',
    'ErrorMessage': 'Check text content',
    'Alert': 'Check visible, get message',
    'Form': 'Fill and submit',
    'List': 'Check items count',
    'Menu': 'Open, select item',
    'Dropdown': 'Open, select option',
    'Avatar': 'Check visible',
    'Badge': 'Get text',
    'Icon': 'Check visible',
    'Image': 'Check loaded',
    'Label': 'Get text',
    'Text': 'Get text content',
    'Tooltip': 'Hover to show',
    'Pagination': 'Click next/prev',
    'Sidebar': 'Check items, click',
    'Navigation': 'Check items, click',
}

# Module definitions
MODULES = {
    "3. Authentication": {
        "routes": ["/login", "/join", "/forgot-password", "/reset-password"],
        "description": "Đăng nhập, đăng ký, quên mật khẩu"
    },
    "4. Dashboard": {
        "routes": ["/dashboard", "/executive/dashboard"],
        "description": "Bảng điều khiển, thống kê tổng quan"
    },
    "5. Tasks": {
        "routes": ["/tasks", "/tasks/[id]", "/tasks/new", "/tasks/kanban"],
        "description": "Quản lý công việc"
    },
    "6. Projects": {
        "routes": ["/projects", "/projects/new", "/projects/[id]"],
        "description": "Danh sách và tạo dự án"
    },
    "7. Project Detail": {
        "routes": [
            "/projects/[id]/overview", "/projects/[id]/gantt", 
            "/projects/[id]/resources", "/projects/[id]/documents",
            "/projects/[id]/import-export", "/projects/[id]/time-locks",
            "/projects/[id]/cost", "/projects/[id]/performance"
        ],
        "description": "Chi tiết dự án: tổng quan, gantt, tài nguyên..."
    },
    "8. Project Settings": {
        "routes": [
            "/projects/[id]/settings/custom-fields",
            "/projects/[id]/settings/field-permissions",
            "/projects/[id]/settings/tags",
            "/projects/[id]/settings/notifications"
        ],
        "description": "Cài đặt dự án"
    },
    "9. Reports": {
        "routes": ["/reports", "/reports/new", "/reports/[id]", 
                   "/reports/cost-analysis", "/reports/performance"],
        "description": "Báo cáo"
    },
    "10. HR Management": {
        "routes": ["/hr/employees", "/hr/employees/[id]", 
                   "/hr/employees/[id]/timeline", "/hr/salary", "/hr/contracts"],
        "description": "Quản lý nhân sự, lương, hợp đồng"
    },
    "11. Time & Activity": {
        "routes": ["/time-logs", "/activity", "/notifications", "/personal-board"],
        "description": "Chấm công, hoạt động, thông báo"
    },
    "12. Settings": {
        "routes": ["/settings/profile", "/settings/workspace", 
                   "/settings/lookups", "/settings/auto-lock", "/settings/recycle-bin"],
        "description": "Cài đặt hệ thống"
    },
    "13. Admin System": {
        "routes": ["/admin/users", "/admin/users/[id]", "/admin/organizations",
                   "/admin/org-approvals", "/admin/deleted-orgs", "/admin/org-recycle-bin",
                   "/admin/quotas", "/admin/audit-logs", "/admin/platform-dashboard",
                   "/admin/impersonation"],
        "description": "Quản trị hệ thống"
    },
    "14. Admin RBAC": {
        "routes": ["/admin/roles", "/admin/roles/customize", 
                   "/admin/members", "/admin/invites"],
        "description": "Quản lý vai trò, quyền hạn"
    },
    "15. Other": {
        "routes": ["/recycle-bin"],
        "description": "Các trang khác"
    }
}

# UI Name mapping
UI_NAME_MAP = {
    "/login": "Đăng nhập",
    "/join": "Tham gia tổ chức",
    "/forgot-password": "Quên mật khẩu",
    "/reset-password": "Đặt lại mật khẩu",
    "/dashboard": "Bảng điều khiển",
    "/executive/dashboard": "Dashboard Giám đốc",
    "/tasks": "Danh sách công việc",
    "/tasks/[id]": "Chi tiết công việc",
    "/tasks/new": "Tạo công việc mới",
    "/tasks/kanban": "Kanban công việc",
    "/projects": "Danh sách dự án",
    "/projects/new": "Tạo dự án mới",
    "/projects/[id]": "Chi tiết dự án - Tasks",
    "/projects/[id]/overview": "Tổng quan dự án",
    "/projects/[id]/gantt": "Biểu đồ Gantt",
    "/projects/[id]/resources": "Tài nguyên dự án",
    "/projects/[id]/documents": "Tài liệu dự án",
    "/projects/[id]/import-export": "Import/Export dự án",
    "/projects/[id]/time-locks": "Khóa kỳ công",
    "/projects/[id]/cost": "Chi phí dự án",
    "/projects/[id]/performance": "Hiệu suất dự án",
    "/projects/[id]/settings/custom-fields": "Trường tùy chỉnh",
    "/projects/[id]/settings/field-permissions": "Phân quyền trường",
    "/projects/[id]/settings/tags": "Quản lý nhãn",
    "/projects/[id]/settings/notifications": "Cài đặt thông báo",
    "/reports": "Danh sách báo cáo",
    "/reports/new": "Tạo báo cáo mới",
    "/reports/[id]": "Chi tiết báo cáo",
    "/reports/cost-analysis": "Phân tích chi phí",
    "/reports/performance": "Báo cáo hiệu suất",
    "/hr/employees": "Danh sách nhân sự",
    "/hr/employees/[id]": "Hồ sơ nhân viên",
    "/hr/employees/[id]/timeline": "Lịch sử làm việc",
    "/hr/salary": "Quản lý lương & cấp bậc",
    "/hr/contracts": "Quản lý hợp đồng",
    "/time-logs": "Nhật ký thời gian",
    "/activity": "Hoạt động gần đây",
    "/notifications": "Thông báo",
    "/personal-board": "Bảng cá nhân",
    "/recycle-bin": "Thùng rác",
    "/settings/profile": "Hồ sơ cá nhân",
    "/settings/workspace": "Cài đặt Workspace",
    "/settings/lookups": "Danh mục hệ thống",
    "/settings/auto-lock": "Khóa tự động",
    "/settings/recycle-bin": "Thùng rác Settings",
    "/admin/users": "Quản lý người dùng",
    "/admin/users/[id]": "Chi tiết người dùng",
    "/admin/organizations": "Quản lý tổ chức",
    "/admin/org-approvals": "Duyệt tổ chức",
    "/admin/deleted-orgs": "Tổ chức đã xóa",
    "/admin/org-recycle-bin": "Thùng rác tổ chức",
    "/admin/members": "Thành viên tổ chức",
    "/admin/invites": "Lời mời đang chờ",
    "/admin/roles": "Quản lý vai trò",
    "/admin/roles/customize": "Tùy chỉnh quyền",
    "/admin/quotas": "Hạn mức hệ thống",
    "/admin/audit-logs": "Nhật ký kiểm toán",
    "/admin/platform-dashboard": "Dashboard Platform",
    "/admin/impersonation": "Impersonation",
}

# User Story mappings from BA documentation
USER_STORY_MAP = {
    "/login": "US-EMP-00-01, US-MNG-00-01, US-CEO-00-01",
    "/join": "US-ORG-01-02",
    "/forgot-password": "US-EMP-00-03",
    "/reset-password": "US-EMP-00-03",
    "/dashboard": "US-EMP-01-01, US-MNG-02-01, US-CEO-01-01",
    "/executive/dashboard": "US-CEO-01-01, US-CEO-01-02, US-CEO-01-03",
    "/tasks": "US-EMP-01-01, US-EMP-01-02",
    "/tasks/[id]": "US-EMP-01-03 ~ US-EMP-01-11",
    "/tasks/new": "US-PM-01-02",
    "/tasks/kanban": "US-EMP-01-01, US-CEO-07-01",
    "/projects": "US-PM-01-01",
    "/projects/new": "US-PM-01-01",
    "/projects/[id]": "US-PM-01-02",
    "/projects/[id]/overview": "US-PM-01-03, US-PM-02-01, US-PM-02-02",
    "/projects/[id]/gantt": "US-PM-02-03",
    "/projects/[id]/resources": "US-PM-03-01, US-PM-03-02",
    "/projects/[id]/documents": "US-PM-03-03",
    "/projects/[id]/import-export": "US-PM-04-01, US-PM-04-02",
    "/projects/[id]/time-locks": "US-PM-05-01, US-PM-05-02",
    "/projects/[id]/cost": "US-MNG-03-01, US-CEO-02-02",
    "/projects/[id]/performance": "US-CEO-01-03",
    "/projects/[id]/settings/custom-fields": "US-PM-06-01",
    "/projects/[id]/settings/field-permissions": "US-PM-06-02",
    "/projects/[id]/settings/tags": "US-PM-06-03",
    "/projects/[id]/settings/notifications": "US-PM-06-04",
    "/reports": "US-MNG-02-02, US-CEO-01-04",
    "/reports/new": "US-MNG-02-02",
    "/reports/[id]": "US-MNG-02-02",
    "/reports/cost-analysis": "US-MNG-03-01, US-CEO-02-02",
    "/reports/performance": "US-CEO-01-03",
    "/hr/employees": "US-ORG-01-01 ~ US-ORG-01-05",
    "/hr/employees/[id]": "US-CEO-01-02, US-ORG-01-03",
    "/hr/employees/[id]/timeline": "US-CEO-01-02",
    "/hr/salary": "US-MNG-03-02, US-CEO-02-01",
    "/hr/contracts": "US-ORG-02-01, US-ORG-02-02",
    "/time-logs": "US-EMP-02-01, US-EMP-02-02, US-EMP-02-03",
    "/activity": "US-EMP-03-01",
    "/notifications": "US-EMP-03-02",
    "/personal-board": "US-EMP-01-01, US-CEO-07-01",
    "/recycle-bin": "US-PM-07-01, US-ORG-03-01",
    "/settings/profile": "US-EMP-00-02",
    "/settings/workspace": "US-ORG-04-01",
    "/settings/lookups": "US-ORG-04-02",
    "/settings/auto-lock": "US-ORG-04-03",
    "/settings/recycle-bin": "US-ORG-03-01",
    "/admin/users": "US-SYS-01-01, US-SYS-01-02, US-SYS-01-03",
    "/admin/users/[id]": "US-SYS-01-02, US-SYS-01-03",
    "/admin/organizations": "US-SYS-02-01, US-SYS-02-02",
    "/admin/org-approvals": "US-SYS-02-03",
    "/admin/deleted-orgs": "US-SYS-02-04",
    "/admin/org-recycle-bin": "US-SYS-02-04",
    "/admin/members": "US-ORG-01-01, US-ORG-01-02",
    "/admin/invites": "US-ORG-01-02",
    "/admin/roles": "US-SYS-03-01, US-ORG-05-01",
    "/admin/roles/customize": "US-SYS-03-02, US-ORG-05-02",
    "/admin/quotas": "US-SYS-04-01",
    "/admin/audit-logs": "US-SYS-05-01",
    "/admin/platform-dashboard": "US-SYS-02-01",
    "/admin/impersonation": "US-SYS-06-01",
}

# Database tables mapping
DATABASE_MAP = {
    "/login": "users, organizations",
    "/join": "users, user_invitations, organizations",
    "/forgot-password": "users, password_reset_tokens",
    "/reset-password": "users, password_reset_tokens",
    "/dashboard": "tasks, projects, time_logs, activity_events",
    "/executive/dashboard": "projects, tasks, users, organizations",
    "/tasks": "tasks, subtasks, task_comments",
    "/tasks/[id]": "tasks, subtasks, task_comments, task_attachments, time_logs",
    "/tasks/new": "tasks, projects, users",
    "/tasks/kanban": "tasks",
    "/projects": "projects, project_members",
    "/projects/new": "projects",
    "/projects/[id]": "projects, tasks, project_members",
    "/projects/[id]/overview": "projects, tasks, project_members, time_logs",
    "/projects/[id]/gantt": "tasks, subtasks",
    "/projects/[id]/resources": "project_resources",
    "/projects/[id]/documents": "project_documents, task_attachments",
    "/projects/[id]/import-export": "tasks, subtasks",
    "/projects/[id]/time-locks": "time_log_locks",
    "/projects/[id]/cost": "time_logs, user_compensations",
    "/projects/[id]/performance": "tasks, time_logs",
    "/projects/[id]/settings/custom-fields": "custom_field_definitions",
    "/projects/[id]/settings/field-permissions": "custom_field_definitions, roles",
    "/projects/[id]/settings/tags": "tags",
    "/projects/[id]/settings/notifications": "notification_settings",
    "/reports": "reports",
    "/reports/new": "reports",
    "/reports/[id]": "reports",
    "/reports/cost-analysis": "time_logs, user_compensations, projects",
    "/reports/performance": "tasks, time_logs, users",
    "/hr/employees": "users, organizations",
    "/hr/employees/[id]": "users, user_compensations, projects",
    "/hr/employees/[id]/timeline": "users, user_compensations, project_members",
    "/hr/salary": "job_levels, user_compensations",
    "/hr/contracts": "contracts, users",
    "/time-logs": "time_logs, tasks, projects",
    "/activity": "activity_events",
    "/notifications": "notifications",
    "/personal-board": "tasks",
    "/recycle-bin": "soft_deleted_items",
    "/settings/profile": "users",
    "/settings/workspace": "organizations",
    "/settings/lookups": "lookup_values",
    "/settings/auto-lock": "auto_lock_settings",
    "/settings/recycle-bin": "soft_deleted_items",
    "/admin/users": "users, roles",
    "/admin/users/[id]": "users, roles",
    "/admin/organizations": "organizations",
    "/admin/org-approvals": "organizations",
    "/admin/deleted-orgs": "organizations",
    "/admin/org-recycle-bin": "organizations",
    "/admin/members": "users, organizations",
    "/admin/invites": "user_invitations",
    "/admin/roles": "roles, permissions",
    "/admin/roles/customize": "roles, permissions, role_permissions",
    "/admin/quotas": "quotas, organizations",
    "/admin/audit-logs": "audit_logs",
    "/admin/platform-dashboard": "organizations, users, projects",
    "/admin/impersonation": "users, audit_logs",
}

# Role access mapping
ROLE_MAP = {
    "/login": "ALL",
    "/join": "PUBLIC",
    "/forgot-password": "PUBLIC",
    "/reset-password": "PUBLIC",
    "/dashboard": "EMP, PM, CEO, ORG_ADMIN",
    "/executive/dashboard": "CEO",
    "/tasks": "EMP, PM, CEO",
    "/tasks/[id]": "EMP, PM, CEO",
    "/tasks/new": "PM",
    "/tasks/kanban": "EMP, PM, CEO",
    "/projects": "PM, CEO, ORG_ADMIN",
    "/projects/new": "PM, ORG_ADMIN",
    "/projects/[id]": "PM, EMP",
    "/projects/[id]/overview": "PM, EMP, CEO",
    "/projects/[id]/gantt": "PM, EMP",
    "/projects/[id]/resources": "PM",
    "/projects/[id]/documents": "PM, EMP",
    "/projects/[id]/import-export": "PM",
    "/projects/[id]/time-locks": "PM",
    "/projects/[id]/cost": "PM, CEO",
    "/projects/[id]/performance": "PM, CEO",
    "/projects/[id]/settings/custom-fields": "PM",
    "/projects/[id]/settings/field-permissions": "PM",
    "/projects/[id]/settings/tags": "PM",
    "/projects/[id]/settings/notifications": "PM",
    "/reports": "PM, CEO, ORG_ADMIN",
    "/reports/new": "PM, CEO",
    "/reports/[id]": "PM, CEO",
    "/reports/cost-analysis": "PM, CEO",
    "/reports/performance": "CEO",
    "/hr/employees": "ORG_ADMIN, CEO",
    "/hr/employees/[id]": "ORG_ADMIN, CEO",
    "/hr/employees/[id]/timeline": "CEO",
    "/hr/salary": "ORG_ADMIN, CEO",
    "/hr/contracts": "ORG_ADMIN, CEO",
    "/time-logs": "EMP, PM, CEO",
    "/activity": "EMP, PM, CEO",
    "/notifications": "EMP, PM, CEO, ORG_ADMIN",
    "/personal-board": "EMP, PM, CEO",
    "/recycle-bin": "PM, ORG_ADMIN",
    "/settings/profile": "EMP, PM, CEO, ORG_ADMIN",
    "/settings/workspace": "ORG_ADMIN",
    "/settings/lookups": "ORG_ADMIN",
    "/settings/auto-lock": "ORG_ADMIN",
    "/settings/recycle-bin": "ORG_ADMIN",
    "/admin/users": "SYS_ADMIN",
    "/admin/users/[id]": "SYS_ADMIN",
    "/admin/organizations": "SYS_ADMIN",
    "/admin/org-approvals": "SYS_ADMIN",
    "/admin/deleted-orgs": "SYS_ADMIN",
    "/admin/org-recycle-bin": "SYS_ADMIN",
    "/admin/members": "ORG_ADMIN",
    "/admin/invites": "ORG_ADMIN",
    "/admin/roles": "SYS_ADMIN, ORG_ADMIN",
    "/admin/roles/customize": "SYS_ADMIN, ORG_ADMIN",
    "/admin/quotas": "SYS_ADMIN",
    "/admin/audit-logs": "SYS_ADMIN",
    "/admin/platform-dashboard": "SYS_ADMIN",
    "/admin/impersonation": "SYS_ADMIN",
}

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
MODULE_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def path_to_route(file_path):
    """Convert file path to route"""
    route = file_path.replace("\\", "/")
    route = route.replace("src/app/(frontend)/", "")
    route = route.replace("/page.tsx", "")
    route = route.replace("(auth)/", "")
    if not route.startswith("/"):
        route = "/" + route
    return route


def detect_element_type(testid):
    """Detect element type from testid name"""
    testid_lower = testid.lower()
    for pattern, elem_type in ELEMENT_TYPES.items():
        if pattern in testid_lower:
            return elem_type
    return "Element"


def generate_description(testid, elem_type):
    """Generate Vietnamese description for testid"""
    testid_clean = testid.replace("-", " ").replace("_", " ")
    
    # Common translations
    translations = {
        "container": "Container chính của trang",
        "page title": "Tiêu đề trang",
        "submit button": "Nút gửi/xác nhận",
        "btn save": "Nút lưu",
        "btn create": "Nút tạo mới",
        "btn delete": "Nút xóa",
        "btn edit": "Nút chỉnh sửa",
        "btn cancel": "Nút hủy",
        "btn close": "Nút đóng",
        "btn add": "Nút thêm",
        "btn refresh": "Nút làm mới",
        "btn export": "Nút xuất dữ liệu",
        "btn import": "Nút nhập dữ liệu",
        "input email": "Ô nhập email",
        "input password": "Ô nhập mật khẩu",
        "input search": "Ô tìm kiếm",
        "input name": "Ô nhập tên",
        "select status": "Dropdown chọn trạng thái",
        "select role": "Dropdown chọn vai trò",
        "filter": "Bộ lọc",
        "table": "Bảng dữ liệu",
        "dialog": "Hộp thoại",
        "modal": "Hộp thoại modal",
        "loading": "Hiệu ứng đang tải",
        "skeleton": "Skeleton loading",
        "empty": "Trạng thái rỗng",
        "error": "Thông báo lỗi",
        "stat": "Thẻ thống kê",
        "card": "Thẻ nội dung",
        "form": "Biểu mẫu",
        "list": "Danh sách",
        "row": "Dòng dữ liệu",
    }
    
    for key, desc in translations.items():
        if key in testid_clean.lower():
            return desc
    
    return f"{elem_type}: {testid_clean}"


def generate_expected_result(testid, elem_type):
    """Generate expected result for testing"""
    results = {
        'Button': "Click được, trigger action",
        'Input': "Nhập được text, validation hoạt động",
        'Select': "Mở dropdown, chọn được option",
        'Filter': "Lọc dữ liệu theo giá trị đã chọn",
        'Checkbox': "Check/uncheck được",
        'Toggle': "Chuyển đổi trạng thái",
        'Table': "Hiển thị dữ liệu, số dòng đúng",
        'TableRow': "Click được, highlight row",
        'Dialog': "Mở/đóng được, content đúng",
        'Container': "Hiển thị đúng khi load trang",
        'Heading': "Text hiển thị đúng",
        'Header': "Hiển thị đúng vị trí",
        'Card': "Hiển thị nội dung đúng",
        'StatCard': "Hiển thị số liệu đúng",
        'Tab': "Chuyển tab, content thay đổi",
        'Link': "Navigate đến đúng trang",
        'Loading': "Hiển thị khi loading, ẩn khi xong",
        'EmptyState': "Hiển thị khi không có dữ liệu",
        'ErrorMessage': "Hiển thị lỗi với message đúng",
        'Alert': "Hiển thị thông báo đúng",
        'Form': "Submit thành công, validation",
        'List': "Hiển thị đúng số item",
        'Pagination': "Chuyển trang, data load đúng",
    }
    return results.get(elem_type, "Hiển thị và hoạt động đúng")


def extract_testids_with_context(file_path):
    """Extract all data-testid values with their context from a file"""
    testids = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Find static testids: data-testid="xxx"
        static_pattern = r'data-testid="([^"]+)"'
        for match in re.finditer(static_pattern, content):
            testid = match.group(1)
            testids.append({
                "testid": testid,
                "is_dynamic": False
            })
        
        # Find dynamic testids: data-testid={`xxx-${var}`}
        dynamic_pattern = r'data-testid=\{[`\']([^`\']+)[`\']\}'
        for match in re.finditer(dynamic_pattern, content):
            testid = match.group(1)
            testids.append({
                "testid": testid,
                "is_dynamic": True
            })
        
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    
    return testids


def get_all_pages():
    """Get all page.tsx files with their testids"""
    pages = {}
    for root, dirs, files in os.walk(FRONTEND_DIR):
        for file in files:
            if file == "page.tsx":
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, BASE_DIR)
                route = path_to_route(rel_path)
                testids = extract_testids_with_context(full_path)
                pages[route] = {
                    "route": route,
                    "source_file": rel_path.replace("\\", "/"),
                    "testids": testids,
                    "testid_count": len(testids),
                    "ui_name": UI_NAME_MAP.get(route, route)
                }
    return pages


def get_module_for_route(route):
    """Get module name for a route"""
    for module_name, module_info in MODULES.items():
        if route in module_info["routes"]:
            return module_name
    return "15. Other"


def create_overview_sheet(wb, pages):
    """Create Sheet 1: UI Checklist overview"""
    ws = wb.active
    ws.title = "1. UI Checklist"
    
    headers = ["STT", "Nhóm/Module", "Tên UI", "Route/URL", "Số TestId", "Source File", "Trạng thái"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    row = 2
    for idx, (route, page) in enumerate(sorted(pages.items()), 1):
        module = get_module_for_route(route)
        ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
        ws.cell(row=row, column=2, value=module).border = THIN_BORDER
        ws.cell(row=row, column=3, value=page["ui_name"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=route).border = THIN_BORDER
        ws.cell(row=row, column=5, value=page["testid_count"]).border = THIN_BORDER
        ws.cell(row=row, column=6, value=page["source_file"]).border = THIN_BORDER
        ws.cell(row=row, column=7, value="Done" if page["testid_count"] >= 5 else "Need Review").border = THIN_BORDER
        
        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 55
    ws.column_dimensions['G'].width = 15


def create_testid_by_file_sheet(wb, pages):
    """Create Sheet 2: TestId by File"""
    ws = wb.create_sheet("2. TestId by File")
    
    headers = ["Source File", "Số lượng", "data-testid List"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    row = 2
    for route, page in sorted(pages.items()):
        ws.cell(row=row, column=1, value=page["source_file"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=page["testid_count"]).border = THIN_BORDER
        testid_list = ", ".join([t["testid"] for t in page["testids"]])
        ws.cell(row=row, column=3, value=testid_list).border = THIN_BORDER
        row += 1
    
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 150


def create_module_sheet(wb, module_name, module_info, pages):
    """Create a detailed sheet for a module"""
    ws = wb.create_sheet(module_name)
    
    # Module header - expand to cover all 16 columns
    ws.merge_cells('A1:P1')
    header_cell = ws.cell(row=1, column=1, value=f"{module_name}: {module_info['description']}")
    header_cell.fill = MODULE_FILL
    header_cell.font = Font(bold=True, color="FFFFFF", size=14)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers - UPDATED with new columns
    headers = [
        "STT", "Page", "Route/URL", "data-testid", "Loại Element", "Mô tả", 
        "User Story ID", "Database Tables", "Role",
        "Hành động Test", "Expected Result", "Priority", 
        "Status", "Bug ID", "Tester", "Ngày test"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # Data rows
    row = 3
    stt = 1
    for route in module_info["routes"]:
        if route not in pages:
            continue
        
        page = pages[route]
        
        # Get additional info from mappings
        user_story = USER_STORY_MAP.get(route, "")
        db_tables = DATABASE_MAP.get(route, "")
        role = ROLE_MAP.get(route, "")
        
        for testid_info in page["testids"]:
            testid = testid_info["testid"]
            elem_type = detect_element_type(testid)
            
            ws.cell(row=row, column=1, value=stt).border = THIN_BORDER
            ws.cell(row=row, column=2, value=page["ui_name"]).border = THIN_BORDER
            ws.cell(row=row, column=3, value=route).border = THIN_BORDER  # Route/URL
            
            # Mark dynamic testids
            testid_display = testid
            if testid_info["is_dynamic"]:
                testid_display = f"[DYNAMIC] {testid}"
            ws.cell(row=row, column=4, value=testid_display).border = THIN_BORDER
            
            ws.cell(row=row, column=5, value=elem_type).border = THIN_BORDER
            ws.cell(row=row, column=6, value=generate_description(testid, elem_type)).border = THIN_BORDER
            ws.cell(row=row, column=7, value=user_story).border = THIN_BORDER  # User Story ID
            ws.cell(row=row, column=8, value=db_tables).border = THIN_BORDER  # Database Tables
            ws.cell(row=row, column=9, value=role).border = THIN_BORDER  # Role
            ws.cell(row=row, column=10, value=ACTION_MAP.get(elem_type, "Check visible")).border = THIN_BORDER
            ws.cell(row=row, column=11, value=generate_expected_result(testid, elem_type)).border = THIN_BORDER
            
            # Priority based on element type
            priority = "High" if elem_type in ['Button', 'Input', 'Form', 'Dialog'] else "Medium"
            ws.cell(row=row, column=12, value=priority).border = THIN_BORDER
            
            ws.cell(row=row, column=13, value="⬜ Chưa test").border = THIN_BORDER
            ws.cell(row=row, column=14, value="").border = THIN_BORDER
            ws.cell(row=row, column=15, value="").border = THIN_BORDER
            ws.cell(row=row, column=16, value="").border = THIN_BORDER
            
            # Alternate row color
            if stt % 2 == 0:
                for col in range(1, 17):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL
            
            row += 1
            stt += 1
    
    # Set column widths - updated for 16 columns
    widths = [5, 22, 28, 35, 12, 35, 45, 40, 25, 20, 30, 10, 15, 10, 15, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header rows
    ws.freeze_panes = 'A3'


def create_statistics_sheet(wb, pages):
    """Create statistics sheet"""
    ws = wb.create_sheet("Statistics")
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="📊 Thống kê UI Checklist")
    title_cell.fill = MODULE_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Stats
    total_pages = len(pages)
    total_testids = sum(p["testid_count"] for p in pages.values())
    
    stats = [
        ("Tổng số Pages", total_pages),
        ("Tổng số data-testid", total_testids),
        ("Trung bình testid/page", round(total_testids / total_pages, 1) if total_pages > 0 else 0),
        ("", ""),
        ("Theo Module:", "Số testid"),
    ]
    
    row = 3
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Module stats
    for module_name, module_info in MODULES.items():
        module_testids = 0
        for route in module_info["routes"]:
            if route in pages:
                module_testids += pages[route]["testid_count"]
        ws.cell(row=row, column=1, value=module_name)
        ws.cell(row=row, column=2, value=module_testids)
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def main():
    """Main function to generate Excel"""
    print("[*] Generating UI Checklist for Testers...")
    
    # Get all pages
    pages = get_all_pages()
    print(f"    Found {len(pages)} pages")
    
    # Create workbook
    wb = Workbook()
    
    # Create sheets
    create_overview_sheet(wb, pages)
    print("    [OK] Created: 1. UI Checklist")
    
    create_testid_by_file_sheet(wb, pages)
    print("    [OK] Created: 2. TestId by File")
    
    # Create module sheets
    for module_name, module_info in MODULES.items():
        create_module_sheet(wb, module_name, module_info, pages)
        print(f"    [OK] Created: {module_name}")
    
    create_statistics_sheet(wb, pages)
    print("    [OK] Created: Statistics")
    
    # Save
    wb.save(OUTPUT_FILE)
    
    total_testids = sum(p["testid_count"] for p in pages.values())
    print(f"\n[DONE] Saved to: {OUTPUT_FILE}")
    print(f"    Total pages: {len(pages)}")
    print(f"    Total data-testid: {total_testids}")
    print(f"    Total sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
