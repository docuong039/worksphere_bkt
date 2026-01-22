#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to generate UI-Story Matrix Excel file.
- 5 sheets for 5 roles: EMP, MNG (PM), CEO, SYS_ADMIN, ORG_ADMIN
- Matrix: X-axis = Epics + Stories, Y-axis = UI Pages + Routes
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(BASE_DIR, "docs", "UI_STORY_MATRIX.xlsx")

# ============= ROLE DEFINITIONS =============
ROLES = {
    "1. EMP": {
        "name": "Nhân viên (Employee)",
        "epics": {
            "Epic EMP-00: Xác thực": ["US-EMP-00-01", "US-EMP-00-02", "US-EMP-00-03"],
            "Epic EMP-01: Quản lý công việc": ["US-EMP-01-01", "US-EMP-01-02", "US-EMP-01-03", "US-EMP-01-04", "US-EMP-01-05", "US-EMP-01-06", "US-EMP-01-07", "US-EMP-01-08", "US-EMP-01-09", "US-EMP-01-10", "US-EMP-01-11"],
            "Epic EMP-02: Ghi nhận thời gian": ["US-EMP-02-01", "US-EMP-02-02", "US-EMP-02-03", "US-EMP-02-04"],
            "Epic EMP-03: Báo cáo định kỳ": ["US-EMP-03-01", "US-EMP-03-02", "US-EMP-03-03"],
            "Epic EMP-04: Activity": ["US-EMP-04-01", "US-EMP-04-02"],
            "Epic EMP-05: Thông báo": ["US-EMP-05-01", "US-EMP-05-02", "US-EMP-05-03", "US-EMP-05-04"],
            "Epic EMP-06: Thùng rác cá nhân": ["US-EMP-06-01", "US-EMP-06-02"],
            "Epic EMP-07: Task cá nhân": ["US-EMP-07-01", "US-EMP-07-02", "US-EMP-07-03", "US-EMP-07-04", "US-EMP-07-05"],
        }
    },
    "2. PM (MNG)": {
        "name": "Quản lý (Project Manager)",
        "epics": {
            "Epic MNG-00: Xác thực": ["US-MNG-00-01", "US-MNG-00-02", "US-MNG-00-03"],
            "Epic MNG-01: Quản lý Dự án": ["US-MNG-01-01", "US-MNG-01-02", "US-MNG-01-03", "US-MNG-01-04", "US-MNG-01-05", "US-MNG-01-06", "US-MNG-01-07", "US-MNG-01-08", "US-MNG-01-09", "US-MNG-01-10", "US-MNG-01-11", "US-MNG-01-12", "US-MNG-01-13", "US-MNG-01-14", "US-MNG-01-15"],
            "Epic MNG-02: Dashboard": ["US-MNG-02-01", "US-MNG-02-02"],
            "Epic MNG-03: Tài nguyên & Chi phí": ["US-MNG-03-01", "US-MNG-03-02", "US-MNG-03-03"],
            "Epic MNG-04: Kiểm soát chu kỳ": ["US-MNG-04-01", "US-MNG-04-02", "US-MNG-04-03"],
            "Epic MNG-05: Tài sản & Quy trình": ["US-MNG-05-01", "US-MNG-05-02", "US-MNG-05-03", "US-MNG-05-04", "US-MNG-05-05"],
            "Epic MNG-06: Activity": ["US-MNG-06-01", "US-MNG-06-02"],
            "Epic MNG-07: Thông báo": ["US-MNG-07-01", "US-MNG-07-02", "US-MNG-07-03"],
            "Epic MNG-08: Thùng rác dự án": ["US-MNG-08-01", "US-MNG-08-02", "US-MNG-08-03"],
            "Epic MNG-09: Biểu đồ Gantt": ["US-MNG-09-01", "US-MNG-09-02", "US-MNG-09-03", "US-MNG-09-04", "US-MNG-09-05"],
            "Epic MNG-10: Task cá nhân": ["US-MNG-10-01", "US-MNG-10-02", "US-MNG-10-03"],
        }
    },
    "3. CEO": {
        "name": "Ban Lãnh đạo (CEO)",
        "epics": {
            "Epic CEO-00: Xác thực": ["US-CEO-00-01", "US-CEO-00-02", "US-CEO-00-03"],
            "Epic CEO-01: Tổng quan Quản trị": ["US-CEO-01-01", "US-CEO-01-02"],
            "Epic CEO-02: Chiến lược Nhân sự": ["US-CEO-02-01", "US-CEO-02-02", "US-CEO-02-03"],
            "Epic CEO-03: Giám sát Báo cáo": ["US-CEO-03-01", "US-CEO-03-02", "US-CEO-03-03"],
            "Epic CEO-04: Activity": ["US-CEO-04-01", "US-CEO-04-02"],
            "Epic CEO-05: Thông báo": ["US-CEO-05-01", "US-CEO-05-02"],
            "Epic CEO-06: Thùng rác công ty": ["US-CEO-06-01", "US-CEO-06-02"],
            "Epic CEO-07: Task cá nhân": ["US-CEO-07-01", "US-CEO-07-02", "US-CEO-07-03"],
        }
    },
    "4. SYS_ADMIN": {
        "name": "Quản trị Hệ thống (System Admin)",
        "epics": {
            "Epic SYS-00: Xác thực": ["US-SYS-00-01", "US-SYS-00-02", "US-SYS-00-03"],
            "Epic SYS-01: Quản trị Nền tảng": ["US-SYS-01-01", "US-SYS-01-02", "US-SYS-01-03", "US-SYS-01-04", "US-SYS-01-05", "US-SYS-01-06", "US-SYS-01-07"],
            "Epic SYS-02: Giám sát & Bảo mật": ["US-SYS-02-01", "US-SYS-02-02"],
            "Epic SYS-03: Thùng rác hệ thống": ["US-SYS-03-01", "US-SYS-03-02", "US-SYS-03-03"],
        }
    },
    "5. ORG_ADMIN": {
        "name": "Quản trị Tổ chức (Org Admin)",
        "epics": {
            "Epic ORG-00: Xác thực": ["US-ORG-00-01", "US-ORG-00-02", "US-ORG-00-03"],
            "Epic ORG-01: Vòng đời Nhân sự": ["US-ORG-01-01", "US-ORG-01-02", "US-ORG-01-03", "US-ORG-01-04", "US-ORG-01-05"],
            "Epic ORG-02: Phân quyền nội bộ": ["US-ORG-02-01", "US-ORG-02-02", "US-ORG-02-03"],
            "Epic ORG-03: Cấu hình Workspace": ["US-ORG-03-01", "US-ORG-03-02", "US-ORG-03-03"],
            "Epic ORG-04: Thùng rác tổ chức": ["US-ORG-04-01", "US-ORG-04-02", "US-ORG-04-03"],
        }
    },
}

# ============= UI PAGES =============
UI_PAGES = [
    # Authentication
    {"name": "Đăng nhập", "route": "/login"},
    {"name": "Tham gia tổ chức", "route": "/join"},
    {"name": "Quên mật khẩu", "route": "/forgot-password"},
    {"name": "Đặt lại mật khẩu", "route": "/reset-password"},
    
    # Dashboard
    {"name": "Bảng điều khiển", "route": "/dashboard"},
    {"name": "Dashboard Giám đốc", "route": "/executive/dashboard"},
    
    # Tasks
    {"name": "Danh sách công việc", "route": "/tasks"},
    {"name": "Chi tiết công việc", "route": "/tasks/[id]"},
    {"name": "Tạo công việc mới", "route": "/tasks/new"},
    {"name": "Kanban công việc", "route": "/tasks/kanban"},
    
    # Projects
    {"name": "Danh sách dự án", "route": "/projects"},
    {"name": "Tạo dự án mới", "route": "/projects/new"},
    {"name": "Chi tiết dự án - Tasks", "route": "/projects/[id]"},
    {"name": "Tổng quan dự án", "route": "/projects/[id]/overview"},
    {"name": "Biểu đồ Gantt", "route": "/projects/[id]/gantt"},
    {"name": "Tài nguyên dự án", "route": "/projects/[id]/resources"},
    {"name": "Tài liệu dự án", "route": "/projects/[id]/documents"},
    {"name": "Import/Export dự án", "route": "/projects/[id]/import-export"},
    {"name": "Khóa kỳ công", "route": "/projects/[id]/time-locks"},
    {"name": "Chi phí dự án", "route": "/projects/[id]/cost"},
    {"name": "Hiệu suất dự án", "route": "/projects/[id]/performance"},
    {"name": "Trường tùy chỉnh", "route": "/projects/[id]/settings/custom-fields"},
    {"name": "Phân quyền trường", "route": "/projects/[id]/settings/field-permissions"},
    {"name": "Quản lý nhãn", "route": "/projects/[id]/settings/tags"},
    {"name": "Cài đặt thông báo", "route": "/projects/[id]/settings/notifications"},
    
    # Reports
    {"name": "Danh sách báo cáo", "route": "/reports"},
    {"name": "Tạo báo cáo mới", "route": "/reports/new"},
    {"name": "Chi tiết báo cáo", "route": "/reports/[id]"},
    {"name": "Phân tích chi phí", "route": "/reports/cost-analysis"},
    {"name": "Báo cáo hiệu suất", "route": "/reports/performance"},
    
    # HR Management
    {"name": "Danh sách nhân sự", "route": "/hr/employees"},
    {"name": "Hồ sơ nhân viên", "route": "/hr/employees/[id]"},
    {"name": "Lịch sử làm việc", "route": "/hr/employees/[id]/timeline"},
    {"name": "Quản lý lương & cấp bậc", "route": "/hr/salary"},
    {"name": "Quản lý hợp đồng", "route": "/hr/contracts"},
    
    # Time & Activity
    {"name": "Nhật ký thời gian", "route": "/time-logs"},
    {"name": "Hoạt động gần đây", "route": "/activity"},
    {"name": "Thông báo", "route": "/notifications"},
    {"name": "Bảng cá nhân", "route": "/personal-board"},
    {"name": "Thùng rác", "route": "/recycle-bin"},
    
    # Settings
    {"name": "Hồ sơ cá nhân", "route": "/settings/profile"},
    {"name": "Cài đặt Workspace", "route": "/settings/workspace"},
    {"name": "Danh mục hệ thống", "route": "/settings/lookups"},
    {"name": "Khóa tự động", "route": "/settings/auto-lock"},
    {"name": "Thùng rác Settings", "route": "/settings/recycle-bin"},
    
    # Admin
    {"name": "Quản lý người dùng", "route": "/admin/users"},
    {"name": "Chi tiết người dùng", "route": "/admin/users/[id]"},
    {"name": "Quản lý tổ chức", "route": "/admin/organizations"},
    {"name": "Duyệt tổ chức", "route": "/admin/org-approvals"},
    {"name": "Tổ chức đã xóa", "route": "/admin/deleted-orgs"},
    {"name": "Thùng rác tổ chức", "route": "/admin/org-recycle-bin"},
    {"name": "Thành viên tổ chức", "route": "/admin/members"},
    {"name": "Lời mời đang chờ", "route": "/admin/invites"},
    {"name": "Quản lý vai trò", "route": "/admin/roles"},
    {"name": "Tùy chỉnh quyền", "route": "/admin/roles/customize"},
    {"name": "Hạn mức hệ thống", "route": "/admin/quotas"},
    {"name": "Nhật ký kiểm toán", "route": "/admin/audit-logs"},
    {"name": "Dashboard Platform", "route": "/admin/platform-dashboard"},
    {"name": "Impersonation", "route": "/admin/impersonation"},
]

# ============= MAPPING UI -> STORIES =============
# This mapping tells which UI page belongs to which stories
UI_STORY_MAPPING = {
    # Authentication
    "/login": ["US-EMP-00-01", "US-MNG-00-01", "US-CEO-00-01", "US-SYS-00-01", "US-ORG-00-01"],
    "/join": ["US-ORG-01-02"],
    "/forgot-password": ["US-EMP-00-03", "US-MNG-00-03", "US-CEO-00-03", "US-SYS-00-03", "US-ORG-00-03"],
    "/reset-password": ["US-EMP-00-03", "US-MNG-00-03", "US-CEO-00-03", "US-SYS-00-03", "US-ORG-00-03"],
    
    # Dashboard
    "/dashboard": ["US-EMP-01-01", "US-MNG-02-01", "US-CEO-01-01"],
    "/executive/dashboard": ["US-CEO-01-01", "US-CEO-01-02", "US-CEO-02-01", "US-CEO-02-02"],
    
    # Tasks
    "/tasks": ["US-EMP-01-01", "US-EMP-01-02"],
    "/tasks/[id]": ["US-EMP-01-03", "US-EMP-01-04", "US-EMP-01-05", "US-EMP-01-06", "US-EMP-01-07", "US-EMP-01-08", "US-EMP-01-09", "US-EMP-01-10", "US-EMP-01-11", "US-EMP-02-01", "US-EMP-02-02", "US-MNG-01-02", "US-MNG-01-03", "US-MNG-01-04", "US-MNG-01-06", "US-MNG-01-07", "US-MNG-01-08", "US-MNG-01-09"],
    "/tasks/new": ["US-MNG-01-02"],
    "/tasks/kanban": ["US-EMP-01-01", "US-EMP-07-02", "US-MNG-10-02", "US-CEO-07-02"],
    
    # Projects
    "/projects": ["US-MNG-01-01", "US-MNG-02-01"],
    "/projects/new": ["US-MNG-01-01"],
    "/projects/[id]": ["US-MNG-01-02", "US-MNG-01-11", "US-MNG-01-14"],
    "/projects/[id]/overview": ["US-MNG-02-01", "US-MNG-02-02", "US-CEO-01-01"],
    "/projects/[id]/gantt": ["US-MNG-09-01", "US-MNG-09-02", "US-MNG-09-03", "US-MNG-09-04", "US-MNG-09-05"],
    "/projects/[id]/resources": ["US-MNG-05-01", "US-MNG-05-02", "US-MNG-05-03", "US-MNG-05-04", "US-MNG-05-05"],
    "/projects/[id]/documents": ["US-MNG-05-01"],
    "/projects/[id]/import-export": ["US-MNG-01-12", "US-MNG-01-15"],
    "/projects/[id]/time-locks": ["US-MNG-04-01", "US-MNG-04-02"],
    "/projects/[id]/cost": ["US-MNG-03-03", "US-CEO-02-02"],
    "/projects/[id]/performance": ["US-MNG-02-02", "US-CEO-01-01"],
    "/projects/[id]/settings/custom-fields": ["US-MNG-01-10"],
    "/projects/[id]/settings/field-permissions": ["US-MNG-01-13"],
    "/projects/[id]/settings/tags": ["US-MNG-01-03"],
    "/projects/[id]/settings/notifications": ["US-MNG-07-03"],
    
    # Reports
    "/reports": ["US-EMP-03-01", "US-EMP-03-02", "US-MNG-04-03", "US-CEO-03-01"],
    "/reports/new": ["US-EMP-03-01"],
    "/reports/[id]": ["US-CEO-03-01", "US-CEO-03-02", "US-CEO-03-03"],
    "/reports/cost-analysis": ["US-MNG-03-03", "US-CEO-02-02"],
    "/reports/performance": ["US-MNG-02-02", "US-CEO-01-01"],
    
    # HR Management
    "/hr/employees": ["US-MNG-03-01", "US-ORG-01-01", "US-ORG-01-02", "US-ORG-01-03", "US-ORG-01-04", "US-ORG-01-05"],
    "/hr/employees/[id]": ["US-MNG-03-01", "US-CEO-01-02", "US-CEO-02-03"],
    "/hr/employees/[id]/timeline": ["US-CEO-01-02"],
    "/hr/salary": ["US-MNG-03-02", "US-CEO-02-01"],
    "/hr/contracts": ["US-CEO-02-03"],
    
    # Time & Activity
    "/time-logs": ["US-EMP-02-01", "US-EMP-02-02", "US-EMP-02-03", "US-EMP-02-04"],
    "/activity": ["US-EMP-04-01", "US-EMP-04-02", "US-MNG-06-01", "US-MNG-06-02", "US-CEO-04-01", "US-CEO-04-02"],
    "/notifications": ["US-EMP-05-01", "US-EMP-05-02", "US-EMP-05-03", "US-EMP-05-04", "US-MNG-07-01", "US-MNG-07-02", "US-CEO-05-01", "US-CEO-05-02"],
    "/personal-board": ["US-EMP-07-01", "US-EMP-07-02", "US-EMP-07-03", "US-EMP-07-04", "US-EMP-07-05", "US-MNG-10-01", "US-MNG-10-02", "US-MNG-10-03", "US-CEO-07-01", "US-CEO-07-02", "US-CEO-07-03"],
    "/recycle-bin": ["US-EMP-06-01", "US-EMP-06-02", "US-MNG-08-01", "US-MNG-08-02", "US-MNG-08-03", "US-CEO-06-01", "US-CEO-06-02", "US-ORG-04-03"],
    
    # Settings
    "/settings/profile": ["US-EMP-00-02", "US-MNG-00-02", "US-CEO-00-02"],
    "/settings/workspace": ["US-ORG-03-01"],
    "/settings/lookups": ["US-ORG-03-03"],
    "/settings/auto-lock": ["US-ORG-03-02"],
    "/settings/recycle-bin": ["US-ORG-04-01", "US-ORG-04-02", "US-ORG-04-03"],
    
    # Admin - System
    "/admin/users": ["US-SYS-01-03", "US-SYS-01-04"],
    "/admin/users/[id]": ["US-SYS-01-04"],
    "/admin/organizations": ["US-SYS-01-01", "US-SYS-01-02"],
    "/admin/org-approvals": ["US-SYS-01-01"],
    "/admin/deleted-orgs": ["US-SYS-03-01"],
    "/admin/org-recycle-bin": ["US-SYS-03-01", "US-SYS-03-02", "US-SYS-03-03"],
    "/admin/members": ["US-ORG-01-01", "US-ORG-01-02", "US-ORG-01-03", "US-ORG-01-04", "US-ORG-02-01", "US-ORG-02-02"],
    "/admin/invites": ["US-ORG-01-02"],
    "/admin/roles": ["US-SYS-01-05", "US-ORG-02-01", "US-ORG-02-02", "US-ORG-02-03"],
    "/admin/roles/customize": ["US-SYS-01-05", "US-ORG-02-03"],
    "/admin/quotas": ["US-SYS-01-07"],
    "/admin/audit-logs": ["US-SYS-02-02"],
    "/admin/platform-dashboard": ["US-SYS-02-01"],
    "/admin/impersonation": ["US-SYS-01-06"],
}

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
EPIC_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
STORY_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
CHECK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")
BLACK_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_role_sheet(wb, sheet_name, role_info):
    """Create a matrix sheet for a role"""
    ws = wb.create_sheet(sheet_name)
    
    epics = role_info["epics"]
    
    # Calculate column positions for each epic
    col_positions = {}  # {story_id: column_index}
    current_col = 3  # Start after UI Name and Route columns
    epic_spans = []  # [(start_col, end_col, epic_name), ...]
    
    for epic_name, stories in epics.items():
        start_col = current_col
        for story in stories:
            col_positions[story] = current_col
            current_col += 1
        end_col = current_col - 1
        epic_spans.append((start_col, end_col, epic_name))
    
    total_cols = current_col - 1
    
    # Row 1: Sheet title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{sheet_name}: {role_info['name']}")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 2: Epic headers (merged cells)
    ws.cell(row=2, column=1, value="Tên giao diện").fill = EPIC_FILL
    ws.cell(row=2, column=1).font = WHITE_FONT
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=2, column=2, value="Route/URL").fill = EPIC_FILL
    ws.cell(row=2, column=2).font = WHITE_FONT
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    for start_col, end_col, epic_name in epic_spans:
        if start_col == end_col:
            cell = ws.cell(row=2, column=start_col, value=epic_name)
        else:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=epic_name)
        cell.fill = EPIC_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # Row 3: Story ID headers
    ws.cell(row=3, column=1, value="").fill = STORY_FILL
    ws.cell(row=3, column=1).border = THIN_BORDER
    ws.cell(row=3, column=2, value="").fill = STORY_FILL
    ws.cell(row=3, column=2).border = THIN_BORDER
    
    for story_id, col_idx in col_positions.items():
        cell = ws.cell(row=3, column=col_idx, value=story_id)
        cell.fill = STORY_FILL
        cell.font = BLACK_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
        cell.border = THIN_BORDER
    
    # Data rows: UI pages
    row = 4
    for idx, page in enumerate(UI_PAGES):
        route = page["route"]
        page_stories = UI_STORY_MAPPING.get(route, [])
        
        # Check if this page has any story from this role
        role_stories = set()
        for epic_name, stories in epics.items():
            role_stories.update(stories)
        
        # Column A: UI Name
        cell_name = ws.cell(row=row, column=1, value=page["name"])
        cell_name.border = THIN_BORDER
        if idx % 2 == 1:
            cell_name.fill = ALT_ROW_FILL
        
        # Column B: Route
        cell_route = ws.cell(row=row, column=2, value=route)
        cell_route.border = THIN_BORDER
        if idx % 2 == 1:
            cell_route.fill = ALT_ROW_FILL
        
        # Check marks for each story
        for story_id, col_idx in col_positions.items():
            cell = ws.cell(row=row, column=col_idx, value="")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if idx % 2 == 1:
                cell.fill = ALT_ROW_FILL
            
            # If this page belongs to this story, mark it
            if story_id in page_stories:
                cell.value = "X"
                cell.fill = CHECK_FILL
                cell.font = Font(bold=True, color="006100")
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    
    for col_idx in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5
    
    # Set row heights
    ws.row_dimensions[2].height = 60  # Epic row
    ws.row_dimensions[3].height = 100  # Story ID row (rotated text)
    
    # Freeze panes
    ws.freeze_panes = 'C4'


def create_summary_sheet(wb):
    """Create a summary sheet"""
    ws = wb.active
    ws.title = "0. Tổng quan"
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="MA TRẬN UI - USER STORY")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Info
    ws.cell(row=3, column=1, value="Mô tả:").font = Font(bold=True)
    ws.cell(row=3, column=2, value="File này thể hiện mối quan hệ giữa các Giao diện (UI) và User Stories theo từng vai trò")
    
    ws.cell(row=5, column=1, value="Hướng dẫn:").font = Font(bold=True)
    ws.cell(row=6, column=1, value="- Trục X (hàng ngang): Tên Epic và các User Story ID")
    ws.cell(row=7, column=1, value="- Trục Y (cột dọc): Tên giao diện và Route/URL")
    ws.cell(row=8, column=1, value="- Ô có dấu 'X': Giao diện đó thuộc User Story tương ứng")
    
    ws.cell(row=10, column=1, value="Các Sheet:").font = Font(bold=True)
    
    row = 11
    for sheet_name, role_info in ROLES.items():
        total_stories = sum(len(stories) for stories in role_info["epics"].values())
        total_epics = len(role_info["epics"])
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=role_info["name"])
        ws.cell(row=row, column=3, value=f"{total_epics} Epics")
        ws.cell(row=row, column=4, value=f"{total_stories} Stories")
        row += 1
    
    ws.cell(row=row + 1, column=1, value="Thống kê:").font = Font(bold=True)
    ws.cell(row=row + 2, column=1, value="Tổng số giao diện:")
    ws.cell(row=row + 2, column=2, value=len(UI_PAGES))
    
    total_stories = sum(len(stories) for role in ROLES.values() for stories in role["epics"].values())
    ws.cell(row=row + 3, column=1, value="Tổng số User Stories:")
    ws.cell(row=row + 3, column=2, value=total_stories)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def main():
    """Main function"""
    print("[*] Generating UI-Story Matrix...")
    
    wb = Workbook()
    
    # Create summary sheet
    create_summary_sheet(wb)
    print("    [OK] Created: 0. Tong quan")
    
    # Create role sheets
    for sheet_name, role_info in ROLES.items():
        create_role_sheet(wb, sheet_name, role_info)
        print(f"    [OK] Created: {sheet_name}")
    
    # Save
    wb.save(OUTPUT_FILE)
    
    print(f"\n[DONE] Saved to: {OUTPUT_FILE}")
    print(f"    Total sheets: {len(wb.sheetnames)}")
    print(f"    Total UI pages: {len(UI_PAGES)}")


if __name__ == "__main__":
    main()
