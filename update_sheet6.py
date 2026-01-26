import openpyxl
from openpyxl.styles import Alignment, Border, Side

excel_path = r'c:\Users\Admin\Documents\worksphere_bkt\docs\UI_STORY_MATRIX.xlsx'
wb = openpyxl.load_workbook(excel_path)

# Current routes in source code and their assigned names
current_routes = [
    ('/activity', 'Hoạt động gần đây'),
    ('/admin/audit-logs', 'Nhật ký kiểm toán'),
    ('/admin/custom-fields', 'Quản lý trường tùy chỉnh toàn cục'),
    ('/admin/impersonation', 'Impersonation'),
    ('/admin/org-approvals', 'Duyệt tổ chức'),
    ('/admin/org-recycle-bin', 'Thùng rác tổ chức'),
    ('/admin/organizations', 'Quản lý tổ chức'),
    ('/admin/platform-dashboard', 'Dashboard Platform'),
    ('/admin/quotas', 'Hạn mức hệ thống'),
    ('/admin/roles', 'Quản lý vai trò'),
    ('/admin/roles/customize', 'Tùy chỉnh quyền'),
    ('/admin/users', 'Quản lý người dùng'),
    ('/admin/users/[id]', 'Chi tiết người dùng'),
    ('/dashboard', 'Bảng điều khiển'),
    ('/executive/dashboard', 'Dashboard Giám đốc'),
    ('/forgot-password', 'Quên mật khẩu'),
    ('/hr/contracts', 'Quản lý hợp đồng'),
    ('/hr/employees', 'Danh sách nhân sự'),
    ('/hr/employees/[id]', 'Hồ sơ nhân viên'),
    ('/hr/employees/[id]/timeline', 'Lịch sử làm việc'),
    ('/hr/salary', 'Quản lý lương & cấp bậc'),
    ('/join', 'Tham gia tổ chức'),
    ('/login', 'Đăng nhập'),
    ('/notifications', 'Thông báo'),
    ('/personal-board', 'Bảng cá nhân'),
    ('/projects', 'Danh sách dự án'),
    ('/projects/[id]', 'Chi tiết dự án - Tasks'),
    ('/projects/[id]/activity', 'Hoạt động dự án'),
    ('/projects/[id]/cost', 'Chi phí dự án'),
    ('/projects/[id]/documents', 'Tài liệu dự án'),
    ('/projects/[id]/gantt', 'Biểu đồ Gantt'),
    ('/projects/[id]/import-export', 'Import/Export dự án'),
    ('/projects/[id]/overview', 'Tổng quan dự án'),
    ('/projects/[id]/performance', 'Hiệu suất dự án'),
    ('/projects/[id]/quality', 'Theo dõi trạng thái test, fix lỗi'),
    ('/projects/[id]/resources', 'Tài nguyên dự án'),
    ('/projects/[id]/settings', 'Cài đặt dự án (Chung)'),
    ('/projects/[id]/settings/custom-fields', 'Trường tùy chỉnh'),
    ('/projects/[id]/settings/field-permissions', 'Phân quyền trường'),
    ('/projects/[id]/settings/lockdown', 'Khóa dự án (Lockdown)'),
    ('/projects/[id]/settings/notifications', 'Cài đặt thông báo'),
    ('/projects/[id]/settings/recycle-bin', 'Thùng rác dự án (Cài đặt)'),
    ('/projects/[id]/settings/tags', 'Quản lý nhãn'),
    ('/projects/[id]/time-locks', 'Khóa kỳ công'),
    ('/projects/[id]/time-logs', 'Nhật ký thời gian dự án'),
    ('/projects/new', 'Tạo dự án mới'),
    ('/recycle-bin', 'Thùng rác'),
    ('/register-org', 'Đăng ký tổ chức mới'),
    ('/reports', 'Danh sách báo cáo'),
    ('/reports/[id]', 'Chi tiết báo cáo'),
    ('/reports/cost-analysis', 'Phân tích chi phí'),
    ('/reports/performance', 'Báo cáo hiệu suất'),
    ('/reset-password', 'Đặt lại mật khẩu'),
    ('/settings/auto-lock', 'Khóa tự động'),
    ('/settings/lookups', 'Danh mục hệ thống'),
    ('/settings/profile', 'Hồ sơ cá nhân'),
    ('/settings/recycle-bin', 'Thùng rác Settings'),
    ('/settings/workspace', 'Cài đặt Workspace'),
    ('/tasks', 'Danh sách công việc'),
    ('/tasks/[id]', 'Chi tiết công việc'),
    ('/tasks/new', 'Tạo công việc mới'),
    ('/time-logs', 'Nhật ký thời gian')
]

# Role abbreviations and their mapping based on logic/common sense for new routes
# Cols: EMP (3), PM (4), CEO (5), SYS (6), ORG (7)
role_access = {
    '/admin/custom-fields': {'SYS': 'X', 'ORG': 'X'},
    '/projects/[id]/activity': {'EMP': 'X', 'PM': 'X', 'CEO': 'X'},
    '/projects/[id]/settings': {'PM': 'X'},
    '/projects/[id]/settings/lockdown': {'PM': 'X'},
    '/projects/[id]/settings/recycle-bin': {'PM': 'X'},
    '/projects/[id]/time-logs': {'EMP': 'X', 'PM': 'X'},
    '/register-org': {'SYS': 'X'}
}

def clean(s):
    if not s: return ""
    return str(s).strip()

if '6. Role Matrix' in wb.sheetnames:
    ws = wb['6. Role Matrix']
    route_col = 2
    start_row = 4
    
    # 1. Find the header row
    found_header = False
    for r in range(1, 10):
        val = clean(ws.cell(row=r, column=route_col).value)
        if val == 'Route/URL':
            start_row = r + 1
            found_header = True
            break
            
    if found_header:
        # 2. Collect routes and mark for deletion
        rows_to_delete = []
        existing_routes = set()
        
        route_list = [r[0] for r in current_routes]
        
        for r in range(start_row, ws.max_row + 1):
            route = clean(ws.cell(row=r, column=route_col).value)
            if not route or route == 'None':
                continue
            if not route.startswith('/'): route = '/' + route
            route = route.replace('\\', '/')
            
            if route not in route_list:
                rows_to_delete.append(r)
            else:
                existing_routes.add(route)
        
        # 3. Delete old rows
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
            
        # 4. Add new routes
        for route, name in current_routes:
            if route not in existing_routes:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=1).value = name
                ws.cell(row=new_row, column=route_col).value = route
                
                # Assign default typical "X" for new routes
                access = role_access.get(route, {})
                if 'EMP' in access: ws.cell(row=new_row, column=3).value = 'X'
                if 'PM' in access: ws.cell(row=new_row, column=4).value = 'X'
                if 'CEO' in access: ws.cell(row=new_row, column=5).value = 'X'
                if 'SYS' in access: ws.cell(row=new_row, column=6).value = 'X'
                if 'ORG' in access: ws.cell(row=new_row, column=7).value = 'X'
                
                # Style
                for c in range(1, 10): # Stylize columns A-I
                    cell = ws.cell(row=new_row, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if c <= 2: cell.alignment = Alignment(horizontal='left', vertical='center')
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Sort the sheet rows by name (optional but nice)
    # Skipping sort for now to keep it simple and safe.

wb.save(excel_path)
print("Sheet 6 updated successfully.")
