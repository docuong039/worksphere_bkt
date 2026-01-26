import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill

excel_path = r'c:\Users\Admin\Documents\worksphere_bkt\docs\UI_STORY_MATRIX.xlsx'
output_path = r'c:\Users\Admin\Documents\worksphere_bkt\docs\UI_STORY_MATRIX_UPDATED.xlsx'

# List of routes currently in source code
current_routes = [
    '/activity', '/admin/audit-logs', '/admin/custom-fields', '/admin/impersonation',
    '/admin/org-approvals', '/admin/org-recycle-bin', '/admin/organizations',
    '/admin/platform-dashboard', '/admin/quotas', '/admin/roles', '/admin/roles/customize',
    '/admin/users', '/admin/users/[id]', '/dashboard', '/executive/dashboard',
    '/forgot-password', '/hr/contracts', '/hr/employees', '/hr/employees/[id]',
    '/hr/employees/[id]/timeline', '/hr/salary', '/join', '/login', '/notifications',
    '/personal-board', '/projects', '/projects/[id]', '/projects/[id]/activity',
    '/projects/[id]/cost', '/projects/[id]/documents', '/projects/[id]/gantt',
    '/projects/[id]/import-export', '/projects/[id]/overview', '/projects/[id]/performance',
    '/projects/[id]/quality', '/projects/[id]/resources', '/projects/[id]/settings',
    '/projects/[id]/settings/custom-fields', '/projects/[id]/settings/field-permissions',
    '/projects/[id]/settings/lockdown', '/projects/[id]/settings/notifications',
    '/projects/[id]/settings/recycle-bin', '/projects/[id]/settings/tags',
    '/projects/[id]/time-locks', '/projects/[id]/time-logs', '/projects/new',
    '/recycle-bin', '/register-org', '/reports', '/reports/[id]', '/reports/cost-analysis',
    '/reports/performance', '/reset-password', '/settings/auto-lock', '/settings/lookups',
    '/settings/profile', '/settings/recycle-bin', '/settings/workspace', '/tasks',
    '/tasks/[id]', '/tasks/new', '/time-logs'
]

# Map of new routes to their names
new_route_names = {
    '/admin/custom-fields': 'Quản lý trường tùy chỉnh toàn cục',
    '/projects/[id]/activity': 'Hoạt động dự án',
    '/projects/[id]/settings': 'Cài đặt dự án (Chung)',
    '/projects/[id]/settings/lockdown': 'Khóa dự án (Lockdown)',
    '/projects/[id]/settings/recycle-bin': 'Thùng rác dự án (Cài đặt)',
    '/projects/[id]/time-logs': 'Nhật ký thời gian dự án',
    '/register-org': 'Đăng ký tổ chức mới'
}

wb = openpyxl.load_workbook(excel_path)

def update_sheet(sheet_name):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    
    # Identify Header row (usually row 2 or 3 has 'Route/URL')
    route_col = 2 # Column B
    start_row = 4 # Data starts here? Judging from previous view
    
    # 1. Find the Route/URL column and data start
    found_header = False
    for r in range(1, 10):
        val = str(ws.cell(row=r, column=route_col).value).strip()
        if val == 'Route/URL':
            start_row = r + 2 # Skip USID row
            found_header = True
            break
    
    if not found_header:
        return

    # 2. Collect routes and mark for deletion
    rows_to_delete = []
    existing_routes = set()
    last_row = ws.max_row
    
    for r in range(start_row, last_row + 1):
        route = str(ws.cell(row=r, column=route_col).value).strip()
        if not route or route == 'None':
            continue
        # Normalize
        if not route.startswith('/'): route = '/' + route
        route = route.replace('\\', '/')
        
        if route not in current_routes:
            rows_to_delete.append(r)
        else:
            existing_routes.add(route)
            
    # 3. Delete rows (in reverse order)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
        
    # 4. Add new routes
    new_to_add = [r for r in current_routes if r not in existing_routes]
    for route in new_to_add:
        name = new_route_names.get(route, route.split('/')[-1].capitalize())
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = name
        ws.cell(row=new_row, column=route_col).value = route
        
        # Style the new row
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=new_row, column=c)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # Add thin border
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Update data sheets
for sheet in ['1. EMP', '2. PM (MNG)', '3. CEO', '4. SYS_ADMIN', '5. ORG_ADMIN']:
    update_sheet(sheet)

# Update Sheet 0 (Summary)
if '0. Tổng quan' in wb.sheetnames:
    ws = wb['0. Tổng quan']
    # Total interfaces is row 22 column 2 (B) roughly
    for r in range(15, 30):
        val = str(ws.cell(row=r, column=1).value)
        if 'Tổng số giao diện' in val:
            ws.cell(row=r, column=2).value = len(current_routes)
            break

wb.save(output_path)
print(f"Updated Excel saved to {output_path}")
print(f"Total current routes: {len(current_routes)}")
